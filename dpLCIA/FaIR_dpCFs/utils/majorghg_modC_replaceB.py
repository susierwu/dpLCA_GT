
"""
Here we compute prospective AGWP / AGTP (and optional CC adjustment) using:
- Scenario- & year-specific ERF (radiative efficiency) per unit concentration (from FaIR, module A)
- Prospective impulse response for decay (pIRF) (from FaIR reruns after a pulse perturbation)

Design philosophy
-----------------
Module B (majorghg_modB_whRF.py) computes metrics using *analytical* decay forms:
- CO2: multi-exponential AR5/AR6-style carbon cycle IRF (a_i, alpha_i)
- CH4/N2O: single exponential with lifetime alpha_gas and then derives AGWP and AGTP in closed form.

Module C replaces the module B analytical decay with externally supplied pIRF(H) for *each gas*. Given pIRF(H) and ERF at the pulse year, we compute:
- RF(H): instantaneous radiative forcing per kg emitted
- AGWP(H): integral of RF from 0..H (numerical)
- AGTP(H): temperature response to RF using the AR6 two-timescale climate response kernel (numerical convolution)
- iAGTP(H): integral of AGTP from 0..H (numerical)

Optionally, we apply the Gasser et al. (2017) carbon-cycle feedback adjustment for CH4/N2O. See notes in `carbon_cycle_adjustment`.

"""

from __future__ import annotations

import numpy as np
from dataclasses import dataclass
from typing import Dict, Optional, Tuple


def _as_float_array(x) -> np.ndarray:
    return np.asarray(x, dtype=float)


def _repeat_H_for_ens(H: np.ndarray, n_ens: int) -> np.ndarray:
    """Return H repeated into shape (len(H), n_ens)."""
    H = _as_float_array(H).reshape(-1)
    return np.repeat(H, n_ens).reshape(-1, n_ens)


def _trapz_cum(y: np.ndarray, dt: float) -> np.ndarray:
    """
    Cumulative trapezoidal integral along axis=0 (time axis).
    y: (T, N)
    Returns: (T, N) with integral from 0..t
    """
    y = _as_float_array(y)
    out = np.zeros_like(y)
    if y.shape[0] <= 1:
        return out
    # trapezoid for each interval
    out[1:] = np.cumsum(0.5 * (y[1:] + y[:-1]) * dt, axis=0)
    return out


def _causal_conv_kernel(rf: np.ndarray, kernel: np.ndarray, dt: float) -> np.ndarray:
    """
    Causal convolution:
      out[t] = ∫_0^t rf[τ] * kernel[t-τ] dτ
    Discrete approximation with dt. Both inputs shape (T, N) for rf and (T,) for kernel.
    Returns out shape (T, N).

    Efficient O(T^2) loop is acceptable for T<=~1000. If going to very large H_max*ts,
    consider FFT with padding.
    """
    rf = _as_float_array(rf)
    k = _as_float_array(kernel).reshape(-1)
    T, N = rf.shape
    if k.shape[0] != T:
        raise ValueError(f"kernel length {k.shape[0]} must match rf length {T}.")
    out = np.zeros_like(rf)
    # loop over lag; vectorized over ensemble dimension
    for lag in range(T):
        # add rf[t-lag] * k[lag] for all t>=lag
        if k[lag] == 0.0:
            continue
        out[lag:] += rf[: T - lag] * (k[lag] * dt)
    return out


@dataclass
class MassConversion:
    """Mass conversion for 1 ppm (CO2) or 1 ppb (CH4/N2O) to kg in atmosphere."""
    M_ATMOS: float = 5.1352e18
    M_AIR: float = 28.97e-3
    M_CO2: float = 44.01e-3
    M_CH4: float = 16.043e-3
    M_N2O: float = 44.0e-3

    def ppm_to_kg_co2(self) -> float:
        return 1e-6 * (self.M_CO2 / self.M_AIR) * self.M_ATMOS

    def ppb_to_kg_ch4(self) -> float:
        return 1e-9 * (self.M_CH4 / self.M_AIR) * self.M_ATMOS

    def ppb_to_kg_n2o(self) -> float:
        return 1e-9 * (self.M_N2O / self.M_AIR) * self.M_ATMOS


class majorghg_metrics_pIRF:
    """
    Compute prospective RF/AGWP/AGTP using pIRF.

    Key inputs supplied (typically from Module A outputs):
    - ERF per unit concentration at each year: erf_diff_t[year_index, ens] in W m-2 (per 1 ppm CO2 or 1 ppb CH4/N2O)
      * For CH4 and N2O, ensure this is **effective** forcing incl. overlap treatment as implemented in the pipeline.
    - pIRF for the pulse year: pirf_H[H_index, ens] as a *dimensionless* decay factor (ΔC(H)/ΔC(0)).
      * Must start at 1.0 at H=0.
      * Must be aligned to the same ensemble dimension as ERF.

    The module does not assume any particular FaIR version; it just consumes arrays.
    """

    def __init__(
        self,
        scn: str,
        H_max: int = 100,
        ts_per_year: int = 1,
        fair_start_y: int = 1750,
        erf_start_y: int = 2000,
        year_index: int = 269,  # 1750+269=2019 default
        mass_conv: Optional[MassConversion] = None,
        # AR6 two-timescale temperature response parameters (used in Module B)
        d: np.ndarray = np.array([3.424102092311, 285.003477841911]),
        q: np.ndarray = np.array([0.443767728883447, 0.313998206372015]),
    ):
        self.scn = scn
        self.H_max = int(H_max)
        self.ts_per_year = int(ts_per_year)
        self.fair_start_y = int(fair_start_y)
        self.erf_start_y = int(erf_start_y)
        self.year_index = int(year_index)

        self.mass = mass_conv or MassConversion()
        self.d = _as_float_array(d).reshape(-1)
        self.q = _as_float_array(q).reshape(-1)
        if self.d.shape != (2,) or self.q.shape != (2,):
            raise ValueError("This implementation assumes the AR6 2-box kernel: d and q must be length-2 arrays.")

        self.dt = 1.0 / self.ts_per_year
        self.H = np.linspace(0, self.H_max, self.H_max * self.ts_per_year + 1)

    # ---------------------------
    # Core metric calculations
    # ---------------------------

    def rf_from_pirf(self, gas: str, erf_diff_t: np.ndarray, pirf_H: np.ndarray,
                     ch4_o3: float = 1.4e-4, ch4_h2o: float = 4.0e-5,
                     n2o_f_n2o_ch4: float = -1.7,  # AR5 8.SM.11.3.3 factor
                     erf_ch4_diff_t: Optional[np.ndarray] = None) -> np.ndarray:
        """
        Compute instantaneous RF(H) per kg emitted using ERF at pulse year and pIRF decay.

        Parameters
        ----------
        gas : {"CO2","CH4","N2O"}
        erf_diff_t : array (T_years, N_ens) effective radiative forcing *difference* for +1 ppm (CO2) or +1 ppb (CH4/N2O)
        pirf_H : array (T_H, N_ens) decay factor ΔC(H)/ΔC(0) for the pulse year
        ch4_o3, ch4_h2o : constant indirect forcing add-ons (W m-2 per 1 ppb) to keep AR6/AR5 style; set to 0 if already included in erf_diff_t
        n2o_f_n2o_ch4 : coefficient for methane destruction effect in N2O RF (AR5 convention)
        erf_ch4_diff_t : required for N2O if including the methane-destruction term via CH4 ERF.

        Returns
        -------
        rf : array (T_H, N_ens) W m-2 per kg emitted
        """
        gas_u = gas.upper()
        erf_diff_t = _as_float_array(erf_diff_t)
        pirf_H = _as_float_array(pirf_H)

        # Map FaIR pulse year index (relative to fair_start_y) to ERF calendar-year index (relative to erf_start_y).
        pulse_year = self.fair_start_y + self.year_index
        erf_idx = pulse_year - self.erf_start_y
        if erf_idx < 0 or erf_idx >= erf_diff_t.shape[0]:
            raise IndexError(
                f"pulse_year={pulse_year} maps to erf_idx={erf_idx}, out of bounds for erf_diff_t with shape {erf_diff_t.shape} "
                f"(erf_start_y={self.erf_start_y})."
            )

        # ensure ensemble dimension matches
        re = erf_diff_t[erf_idx]  # (N_ens,)
        if pirf_H.shape[1] != re.shape[0]:
            raise ValueError(f"Ensemble mismatch: pirf_H has {pirf_H.shape[1]} ens but ERF has {re.shape[0]}.")

        if pirf_H.shape[0] != self.H.shape[0]:
            raise ValueError(f"H length mismatch: pirf_H has {pirf_H.shape[0]} rows but expected {self.H.shape[0]}.")

        # concentration->kg conversion
        if gas_u == "CO2":
            kg_per_unit = self.mass.ppm_to_kg_co2()
            A = re / kg_per_unit  # W m-2 per kg at H=0
            rf = pirf_H * A  # (T_H, N)
            return rf

        if gas_u == "CH4":
            kg_per_unit = self.mass.ppb_to_kg_ch4()
            # If erf_diff_t already includes indirect forcing, set ch4_o3/ch4_h2o to 0.
            A = (re + ch4_o3 + ch4_h2o) / kg_per_unit
            rf = pirf_H * A
            return rf

        if gas_u == "N2O":
            if erf_ch4_diff_t is None:
                raise ValueError("For N2O RF including methane-destruction term, provide erf_ch4_diff_t.")
            erf_ch4_diff_t = _as_float_array(erf_ch4_diff_t)
            re_ch4 = erf_ch4_diff_t[erf_idx]
            kg_per_unit = self.mass.ppb_to_kg_n2o()
            # Include methane-destruction term (AR5). If N2O ERF already includes this,
            # set n2o_f_n2o_ch4=0 and omit erf_ch4_diff_t.
            A = (re + n2o_f_n2o_ch4 * re_ch4) / kg_per_unit
            rf = pirf_H * A
            return rf

        raise ValueError("gas must be one of {'CO2','CH4','N2O'}.")

    def agwp_from_rf(self, rf: np.ndarray) -> np.ndarray:
        """AGWP(H) = ∫ RF(t) dt, cumulative trapezoid."""
        return _trapz_cum(rf, self.dt)

    def agtp_from_rf(self, rf: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        """
        AGTP(H) from generic RF(H) using AR6 2-box temperature response kernel.

        Kernel for each box j:
          k_j(s) = (q_j / d_j) * exp(-s / d_j)
        Then:
          AGTP(t) = ∫_0^t RF(τ) * Σ_j k_j(t-τ) dτ

        Returns
        -------
        agtp : (T_H, N_ens)
        iagtp : (T_H, N_ens) = ∫_0^t AGTP(s) ds
        """
        rf = _as_float_array(rf)
        T, N = rf.shape
        # build time-lag vector in years
        s = self.H  # length T

        agtp = np.zeros_like(rf)
        for j in range(2):
            k = (self.q[j] / self.d[j]) * np.exp(-s / self.d[j])  # length T
            agtp += _causal_conv_kernel(rf, k, self.dt)

        iagtp = _trapz_cum(agtp, self.dt)
        return agtp, iagtp

    # ---------------------------
    # Carbon-cycle feedback adjustment (optional)
    # ---------------------------

    def carbon_cycle_adjustment(
        self,
        agtp_gas: np.ndarray,
        rf_co2: np.ndarray,
        agwp_co2: np.ndarray,
        *,
        gamma: float = 3.015e12,  # kgCO2/yr/K (Gasser et al. 2017)
        a_gasser: np.ndarray = np.array([0.6368, 0.3322, 0.0310]),
        alpha_gasser: np.ndarray = np.array([2.376, 30.14, 490.1]),
        M_CO2: float = 44.01e-3,
        M_C: float = 12.0e-3,
    ) -> Tuple[np.ndarray, np.ndarray]:
        """
        Carbon-cycle feedback adjustment used in Module B (Gasser et al., 2017 approach).

        What it represents:
        - A warming-driven additional CO2 flux (kgCO2/yr) proportional to AGTP_gas via gamma,
          then converted into an added CO2 forcing and AGWP by convolving with CO2 RF/AGWP.

        Need for this with pIRF:
        - If pIRF for CH4/N2O comes from rerunning FaIR and FaIR is configured such that
          *non-CO2 warming induces extra CO2 emissions/uptake anomalies* (a full climate–carbon feedback),
          then this adjustment could be double-counting.
        - In standard FaIR setups, temperature affects the CO2 impulse response (iIRF100), but
          that changes how CO2 emissions decay; it does **not** necessarily generate a new CO2 source term
          solely because CH4 warmed the climate. In that common case, this adjustment is still a distinct effect.

        Implementation notes:
        - This function returns (rf_cc, agwp_cc). You add them to the gas's own RF/AGWP.
        - The temperature feedback part (agtp_cc) is typically small and often omitted; Module B returns agtp_cc
          but kept it as zeros. Here we return only rf_cc and agwp_cc to avoid implying a full derivation.

        Shapes
        ------
        agtp_gas, rf_co2, agwp_co2: (T_H, N_ens)

        Returns
        -------
        rf_cc : (T_H, N_ens)
        agwp_cc : (T_H, N_ens)
        """
        agtp_gas = _as_float_array(agtp_gas)
        rf_co2 = _as_float_array(rf_co2)
        agwp_co2 = _as_float_array(agwp_co2)

        if agtp_gas.shape != rf_co2.shape or agtp_gas.shape != agwp_co2.shape:
            raise ValueError("All inputs must have the same shape (T_H, N_ens).")

        T, N = agtp_gas.shape
        s = self.H  # length T
        dt = self.dt

        # Build r_f kernel from Gasser (discrete version of Module B)
        a_gasser = _as_float_array(a_gasser).reshape(-1)
        alpha_gasser = _as_float_array(alpha_gasser).reshape(-1)
        if a_gasser.shape != (3,) or alpha_gasser.shape != (3,):
            raise ValueError("a_gasser and alpha_gasser must be length-3 arrays.")

        # Discrete-time r_f(s) (note: Module B used an ad-hoc r_f[0]=sum(a)/dts; keep consistent)
        r_f = np.zeros(T, dtype=float)
        r_f[0] = np.sum(a_gasser) / dt
        for i in range(3):
            r_f -= (a_gasser[i] / alpha_gasser[i]) * np.exp(-s / alpha_gasser[i])

        # 1) Warming-driven CO2 flux (kgCO2/yr) from AGTP_gas
        # F_CO2(t) = ∫_0^t AGTP_gas(τ) * gamma * r_f(t-τ) dτ
        # Here r_f is treated as a kernel; we convolve along time with dt.
        F_CO2 = _causal_conv_kernel(agtp_gas * gamma, r_f, dt)  # (T, N)

        # 2) Added CO2 RF and AGWP from that flux, convolving with CO2 rf/agwp
        rf_cc = np.zeros_like(rf_co2)
        agwp_cc = np.zeros_like(agwp_co2)

        # Convert kgCO2/yr into kgC/yr scaling used in Module B via (M_CO2/M_C)
        scale = (M_CO2 / M_C)

        # The line above is intentionally strict: convolution kernel must be 1D.
        # But rf_co2 differs by ensemble member. To preserve ensemble dependence, do lag-loop manually:
        rf_cc = np.zeros_like(rf_co2)
        agwp_cc = np.zeros_like(agwp_co2)
        for lag in range(T):
            if lag == 0:
                k_rf = rf_co2[0]  # (N,)
                k_agwp = agwp_co2[0]
            else:
                k_rf = rf_co2[lag]
                k_agwp = agwp_co2[lag]
            rf_cc[lag:] += (F_CO2[: T - lag] * scale) * (k_rf * dt)
            agwp_cc[lag:] += (F_CO2[: T - lag] * scale) * (k_agwp * dt)

        return rf_cc, agwp_cc

    # ---------------------------
    # Convenience wrappers per gas
    # ---------------------------

    def compute_metrics_for_gas(
        self,
        gas: str,
        erf_diff_t: np.ndarray,
        pirf_H: np.ndarray,
        *,
        apply_cc: bool = False,
        # CC adjustment needs CO2 metrics:
        rf_co2: Optional[np.ndarray] = None,
        agwp_co2: Optional[np.ndarray] = None,
        # CH4/N2O extras:
        erf_ch4_diff_t: Optional[np.ndarray] = None,
        ch4_o3: float = 1.4e-4,
        ch4_h2o: float = 4.0e-5,
        n2o_f_n2o_ch4: float = -1.7,
    ) -> Dict[str, np.ndarray]:
        """
        Returns dict with keys: rf, agwp, agtp, iagtp, and if apply_cc: rf_cc, agwp_cc, rf_final, agwp_final, gwp_final.
        GWP_final is AGWP(H)/AGWP_CO2(H), with year 0 handled as 0.

        Notes:
        - For CO2, apply_cc is ignored.
        - For CH4/N2O, apply_cc requires rf_co2 and agwp_co2 (computed for the same H grid).
        """
        gas_u = gas.upper()

        rf = self.rf_from_pirf(
            gas_u,
            erf_diff_t=erf_diff_t,
            pirf_H=pirf_H,
            ch4_o3=ch4_o3,
            ch4_h2o=ch4_h2o,
            n2o_f_n2o_ch4=n2o_f_n2o_ch4,
            erf_ch4_diff_t=erf_ch4_diff_t,
        )
        agwp = self.agwp_from_rf(rf)
        agtp, iagtp = self.agtp_from_rf(rf)

        out = {"rf": rf, "agwp": agwp, "agtp": agtp, "iagtp": iagtp}

        if gas_u == "CO2":
            # by definition, GWP_CO2 = 1; provide it for symmetry
            gwp = np.ones_like(agwp)
            gwp[0] = 1.0
            out["gwp"] = gwp
            return out

        if apply_cc:
            if rf_co2 is None or agwp_co2 is None:
                raise ValueError("apply_cc=True requires rf_co2 and agwp_co2 arrays.")
            rf_cc, agwp_cc = self.carbon_cycle_adjustment(agtp_gas=agtp, rf_co2=rf_co2, agwp_co2=agwp_co2)
            rf_final = rf + rf_cc
            agwp_final = agwp + agwp_cc

            gwp_final = np.zeros_like(agwp_final)
            # avoid divide by zero at H=0
            gwp_final[1:] = agwp_final[1:] / agwp_co2[1:]

            out.update({
                "rf_cc": rf_cc,
                "agwp_cc": agwp_cc,
                "rf_final": rf_final,
                "agwp_final": agwp_final,
                "gwp_final": gwp_final,
            })
        else:
            gwp = np.zeros_like(agwp)
            gwp[1:] = agwp[1:] / agwp_co2[1:] if agwp_co2 is not None else np.nan
            out["gwp"] = gwp

        return out


# =========================
# I/O helpers (Module C)
# =========================

import os
import pickle
from pathlib import Path
from typing import Any


def load_pirf_ens_pickle(pkl_path: str) -> Dict[str, Dict[str, np.ndarray]]:
    """Load pIRF ensemble pickle.

    Expected structure (as in ms runs):
        pirf[scn][gas] -> np.ndarray shape (T_H, N_ens)

    Returns:
        dict: {scn: {gas: pirf_H}}
    """
    with open(pkl_path, "rb") as f:
        obj = pickle.load(f)
    if not isinstance(obj, dict):
        raise TypeError(f"pIRF pickle must be dict, got {type(obj)}")
    return obj


def load_erf_ens_pickle(pkl_path: str) -> Dict[str, Dict[str, np.ndarray]]:
    """Load ERF/lifetime pickle produced by the ERF_C pipeline.

    Expected structure (observed in ERF_C_lifetime_CO2CH4N2O_1001ensemble_dicts*.pkl):
        erf['co2_all_ensemble_dict']['gas_re'][scn] -> np.ndarray (T_years, N_ens)
        erf['ch4_all_ensemble_dict']['gas_re'][scn] -> np.ndarray (T_years, N_ens)
        erf['n2o_all_ensemble_dict']['gas_re'][scn] -> np.ndarray (T_years, N_ens)

    Returns:
        dict: {gas: {scn: erf_diff_t}}
        where gas keys are 'CO2','CH4','N2O'
    """
    with open(pkl_path, "rb") as f:
        obj = pickle.load(f)
    if not isinstance(obj, dict):
        raise TypeError(f"ERF pickle must be dict, got {type(obj)}")

    def _pull(gkey: str) -> Dict[str, np.ndarray]:
        if gkey not in obj:
            raise KeyError(f"Missing key {gkey} in ERF pickle")
        d = obj[gkey]
        if not isinstance(d, dict) or "gas_re" not in d:
            raise KeyError(f"ERF pickle[{gkey}] must be dict with 'gas_re'")
        re_dict = d["gas_re"]
        if not isinstance(re_dict, dict):
            raise TypeError(f"ERF pickle[{gkey}]['gas_re'] must be dict")
        return re_dict

    out = {
        "CO2": _pull("co2_all_ensemble_dict"),
        "CH4": _pull("ch4_all_ensemble_dict"),
        "N2O": _pull("n2o_all_ensemble_dict"),
    }
    return out


def _write_matrix_sheet(ws, name: str, H: np.ndarray, mat: np.ndarray) -> None:
    """Write matrix (T, N) with header row = ensemble index, col A = H."""
    ws.title = name
    T, N = mat.shape
    # header
    ws.cell(1, 1, value=None)
    for j in range(N):
        ws.cell(1, 2 + j, value=j)
    # body
    for i in range(T):
        ws.cell(2 + i, 1, value=float(H[i]))
        for j in range(N):
            ws.cell(2 + i, 2 + j, value=float(mat[i, j]))


def _write_point_sheet(ws, name: str, H: np.ndarray, vec: np.ndarray, gas_label: str) -> None:
    """Write pointvalue vector (T,) with 2 columns: H and value; top row has gas label."""
    ws.title = name
    ws.cell(1, 1, value=None)
    ws.cell(1, 2, value=gas_label)
    for i in range(len(H)):
        ws.cell(2 + i, 1, value=float(H[i]))
        ws.cell(2 + i, 2, value=float(vec[i]))


def save_metrics_excel(
    out_xlsx: str,
    gas: str,
    model_year: int,
    H: np.ndarray,
    metrics: Dict[str, np.ndarray],
    *,
    include_gwp_sheet: bool = True,
    point_stat: str = "median",
) -> None:
    """Save compute_metrics_for_gas() outputs to an Excel like Module B (without dcf).

    Sheets created (ensemble + pointvalue):
      - rf_wh_ensmb_{gas}{MY}
      - agwp_wh_ensmb_{gas}{MY}
      - agtp_wh_ensmb_{gas}{MY}
      - iagtp_wh_ensmb_{gas}{MY}
      - (optional) gwp_wh_ensmb_{gas}{MY}   (if metrics provides 'gwp_final' or 'gwp')

    Pointvalue sheets:
      - rf_pointvalue{gas}_{MY}
      - agwp_pointvalue{gas}_{MY}
      - agtp_pointvalue{gas}_{MY}
      - iagtp_pointvalue{gas}_{MY}
      - (optional) gwp_pointvalue{gas}_{MY}

    If apply_cc was used, and keys exist, we also write:
      - rf_cc_wh_ensmb_{gas}{MY}, agwp_cc_wh_ensmb_{gas}{MY}
      - rf_final_wh_ensmb_{gas}{MY}, agwp_final_wh_ensmb_{gas}{MY}
      - and pointvalue versions for each.

    point_stat: 'median' or 'mean'
    """
    from openpyxl import Workbook

    gas_u = gas.upper()
    MY = int(model_year)

    def _point(mat: np.ndarray) -> np.ndarray:
        if mat.ndim == 1:
            return mat
        if point_stat == "mean":
            return np.mean(mat, axis=1)
        return np.median(mat, axis=1)

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    core_keys = ["rf", "agwp", "agtp", "iagtp"]
    for k in core_keys:
        if k not in metrics:
            raise KeyError(f"metrics missing key '{k}'")
        mat = _as_float_array(metrics[k])
        if mat.ndim == 1:
            mat = mat.reshape(-1, 1)
        ens_name = f"{k}_wh_ensmb_{gas_u}{MY}"
        pt_name = f"{k}_pointvalue{gas_u}_{MY}"
        _write_matrix_sheet(wb.create_sheet(), ens_name, H, mat)
        _write_point_sheet(wb.create_sheet(), pt_name, H, _point(mat), gas_u)

    # Optional GWP sheet: accept either 'gwp_final' (apply_cc case) or 'gwp'
    if include_gwp_sheet:
        gkey = None
        for cand in ["gwp_final", "gwp"]:
            if cand in metrics:
                gkey = cand
                break
        if gkey is not None:
            mat = _as_float_array(metrics[gkey])
            if mat.ndim == 1:
                mat = mat.reshape(-1, 1)
            ens_name = f"gwp_wh_ensmb_{gas_u}{MY}"
            pt_name = f"gwp_pointvalue{gas_u}_{MY}"
            _write_matrix_sheet(wb.create_sheet(), ens_name, H, mat)
            _write_point_sheet(wb.create_sheet(), pt_name, H, _point(mat), gas_u)

    # Carbon-cycle related extras (write if present)
    extra_keys = ["rf_cc", "agwp_cc", "rf_final", "agwp_final"]
    for k in extra_keys:
        if k in metrics:
            mat = _as_float_array(metrics[k])
            if mat.ndim == 1:
                mat = mat.reshape(-1, 1)
            ens_name = f"{k}_wh_ensmb_{gas_u}{MY}"
            pt_name = f"{k}_pointvalue{gas_u}_{MY}"
            _write_matrix_sheet(wb.create_sheet(), ens_name, H, mat)
            _write_point_sheet(wb.create_sheet(), pt_name, H, _point(mat), gas_u)

    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def compute_and_save_case(
    *,
    scn: str,
    model_year: int,
    fair_start_y: int,
    erf_start_y: int,
    year_index: int,
    H_max: int,
    ts_per_year: int,
    pirf_pkl: str,
    erf_pkl: str,
    out_dir: str,
    apply_cc: bool = False,
    save_pickle: bool = True,
    pickle_basename: Optional[str] = None,
) -> Dict[str, Dict[str, np.ndarray]]:
    """One-stop helper for one (scenario, model_year) case.

    - loads pIRF and ERF pickles
    - computes CO2 first, then CH4 and N2O (using CO2 as GWP denominator)
    - writes 3 Excel files (CO2/CH4/N2O) into out_dir using naming rule:
        agwp_by_pIRF_tstep{ts}GAS_{scn}_fair_start{fair_start_y}MY{model_year}.xlsx

    Returns:
        dict: {'CO2': metrics, 'CH4': metrics, 'N2O': metrics}
    """
    pirf_all = load_pirf_ens_pickle(pirf_pkl)
    erf_all = load_erf_ens_pickle(erf_pkl)

    scn_l = scn.lower()
    if scn_l not in pirf_all:
        raise KeyError(f"Scenario {scn} not found in pIRF pickle")
    if scn_l not in erf_all["CO2"]:
        raise KeyError(f"Scenario {scn} not found in ERF pickle")

    mc = majorghg_metrics_pIRF(
        scn=scn_l,
        H_max=H_max,
        ts_per_year=ts_per_year,
        fair_start_y=fair_start_y,
        erf_start_y=erf_start_y,
        year_index=year_index,
    )

    pirf_co2 = pirf_all[scn_l]["CO2"]
    pirf_ch4 = pirf_all[scn_l]["CH4"]
    pirf_n2o = pirf_all[scn_l]["N2O"]

    erf_co2 = erf_all["CO2"][scn_l]
    erf_ch4 = erf_all["CH4"][scn_l]
    erf_n2o = erf_all["N2O"][scn_l]

    co2 = mc.compute_metrics_for_gas("CO2", erf_co2, pirf_co2)

    ch4 = mc.compute_metrics_for_gas(
        "CH4", erf_ch4, pirf_ch4,
        agwp_co2=co2["agwp"],
        apply_cc=apply_cc,
        rf_co2=co2.get("rf"),
    )

    n2o = mc.compute_metrics_for_gas(
        "N2O", erf_n2o, pirf_n2o,
        agwp_co2=co2["agwp"],
        apply_cc=apply_cc,
        rf_co2=co2.get("rf"),
        erf_ch4_diff_t=erf_ch4,
    )

    H = mc.H

    def _fname(gas: str) -> str:
        return os.path.join(
            out_dir,
            f"agwp_by_pIRF_tstep{ts_per_year}{gas}_{scn_l}_fair_start{fair_start_y}MY{int(model_year)}.xlsx",
        )

    save_metrics_excel(_fname("CO2"), "CO2", model_year, H, co2)
    save_metrics_excel(_fname("CH4"), "CH4", model_year, H, ch4)
    save_metrics_excel(_fname("N2O"), "N2O", model_year, H, n2o)

    results = {"CO2": co2, "CH4": ch4, "N2O": n2o}

    if save_pickle:
        if pickle_basename is None:
            pickle_basename = (
                f"metrics_by_pIRF_tstep{ts_per_year}ALL_{scn_l}_fair_start{fair_start_y}MY{int(model_year)}.pkl"
            )
        pkl_path = os.path.join(out_dir, pickle_basename)
        payload = {
            "meta": {
                "scn": scn_l,
                "model_year": int(model_year),
                "fair_start_y": int(fair_start_y),
                "erf_start_y": int(erf_start_y),
                "year_index": int(year_index),
                "H_max": int(H_max),
                "ts_per_year": int(ts_per_year),
                "H": H,
            },
            "metrics": results,
        }
        with open(pkl_path, "wb") as f:
            pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)

    return results